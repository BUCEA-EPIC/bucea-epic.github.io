import os
import json
import io
from datetime import datetime
from urllib.parse import quote
import openpyxl
from openpyxl.styles import Font, Alignment

from fastapi import APIRouter, Form, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from app.core.database import get_db
from app.core.config import UPLOAD_DIR, ADMIN_KEY, logger
from app.core.security import create_access_token, verify_admin
from app.models.tables import Submission, Student

router = APIRouter(prefix="/admin")

# --- 🛠️ 内部工具函数：生成 Excel 文件流 ---
def _generate_excel_common(db: Session):
    # 1. 查询所有数据
    submissions = db.query(Submission).options(
        joinedload(Submission.students)
    ).order_by(Submission.submission_time.desc()).all()

    # 2. 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "比赛提交记录"

    # 表头
    headers = [
        "ID", "赛道名称", "成员信息 (姓名-班级-学号)", 
        "提交时间", "原始文件名", "评分", "是否收藏", "是否已评", "备注"
    ]
    ws.append(headers)

    # 样式设置
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 3. 填充数据
    for sub in submissions:
        # 格式化学生信息 (换行显示)
        students_info = "\n".join([
            f"{s.name} - {s.cls} - {s.student_id}" 
            for s in sub.students
        ])
        
        submit_time_str = sub.submission_time.strftime("%Y-%m-%d %H:%M:%S") if sub.submission_time else ""

        row = [
            sub.id,
            sub.track_name,
            students_info,
            submit_time_str,
            sub.original_filename,
            sub.score,
            "是" if sub.is_starred else "否",
            "是" if sub.is_graded else "否",
            sub.remark
        ]
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['I'].width = 30
    
    # 自动换行
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # 4. 导出到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# --- 管理员后台导出 ---
@router.get("/export_excel", dependencies=[Depends(verify_admin)])
async def export_excel(db: Session = Depends(get_db)):
    output = _generate_excel_common(db)
    
    filename = f"比赛统计_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    encoded_filename = quote(filename) # 处理中文文件名
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"}
    )

# ==========================================
# 管理接口
# ==========================================

@router.post("/login")
async def admin_login(key: str = Form(...), db: Session = Depends(get_db)):
    if not ADMIN_KEY:
        logger.error("管理员密钥未配置")
        raise HTTPException(status_code=500, detail="管理员密钥未配置")

    if key != ADMIN_KEY:
        logger.warning("⚠️ 登录失败，密钥错误")
        raise HTTPException(status_code=401, detail="密钥错误")

    token = create_access_token(db)
    logger.info("🔑 管理员登录成功")
    return {"status": "success", "token": token}

@router.get("/submissions", dependencies=[Depends(verify_admin)])
async def get_submissions(db: Session = Depends(get_db)):
    submissions = db.query(Submission).options(
        joinedload(Submission.students)
    ).order_by(Submission.submission_time.desc()).all()
    return submissions

@router.delete("/submission", dependencies=[Depends(verify_admin)])
async def delete_submission(id: int, db: Session = Depends(get_db)):
    sub = db.query(Submission).filter(Submission.id == id).first()
    if not sub:
        raise HTTPException(status_code=404, detail="未找到")

    if sub.filename:
        try:
            os.remove(os.path.join(UPLOAD_DIR, sub.filename))
        except OSError as exc:
            logger.warning("删除上传文件失败: %s", exc)

    db.delete(sub)
    db.commit()
    return {"status": "success"}

@router.post("/update_submission", dependencies=[Depends(verify_admin)])
async def update_submission(
    id: int = Form(...),
    track_name: str = Form(...),
    student_infos: str = Form(...),
    # target_email 已移除
    db: Session = Depends(get_db)
):
    sub = db.query(Submission).filter(Submission.id == id).first()
    if not sub:
        raise HTTPException(status_code=404, detail="未找到")

    sub.track_name = track_name

    db.query(Student).filter(Student.submission_id == id).delete()

    try:
        new_students_data = json.loads(student_infos)
        for s in new_students_data:
            db.add(Student(
                submission_id=id,
                name=s.get('name'),
                cls=s.get('cls') or s.get('className'),
                student_id=s.get('id') or s.get('studentId')
            ))
    except Exception as e:
        logger.error(f"更新解析失败: {e}")
        raise HTTPException(status_code=422, detail="数据格式错误")

    db.commit()
    return {"status": "success"}

@router.post("/toggle_star", dependencies=[Depends(verify_admin)])
async def toggle_star(id: int = Form(...), is_starred: bool = Form(...), db: Session = Depends(get_db)):
    sub = db.query(Submission).filter(Submission.id == id).first()
    if sub:
        sub.is_starred = is_starred
        db.commit()
    return {"status": "success"}

@router.post("/update_score", dependencies=[Depends(verify_admin)])
async def update_score(id: int = Form(...), score: int = Form(...), db: Session = Depends(get_db)):
    sub = db.query(Submission).filter(Submission.id == id).first()
    if sub:
        sub.score = score
        sub.is_graded = True
        db.commit()
    return {"status": "success"}

@router.post("/toggle_graded", dependencies=[Depends(verify_admin)])
async def toggle_graded(id: int = Form(...), is_graded: bool = Form(...), db: Session = Depends(get_db)):
    sub = db.query(Submission).filter(Submission.id == id).first()
    if sub:
        sub.is_graded = is_graded
        db.commit()
    return {"status": "success"}

@router.post("/update_remark", dependencies=[Depends(verify_admin)])
async def update_remark(id: int = Form(...), remark: str = Form(...), db: Session = Depends(get_db)):
    sub = db.query(Submission).filter(Submission.id == id).first()
    if sub:
        sub.remark = remark
        db.commit()
    return {"status": "success"}
